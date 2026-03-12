"""
XLSX / CSV → Text converter.

Reads spreadsheet data and yields text one sheet (or chunk) at a time.
"""

import csv
import logging
from typing import Generator

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

ROWS_PER_CHUNK = 200


def _xlsx_to_text(file_path: str) -> Generator[str, None, None]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        buffer: list[str] = []
        buffer.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            buffer.append(" | ".join(cells))
            if len(buffer) >= ROWS_PER_CHUNK:
                yield "\n".join(buffer) + "\n"
                buffer.clear()
        if buffer:
            yield "\n".join(buffer) + "\n"
    wb.close()


def _csv_to_text(file_path: str) -> Generator[str, None, None]:
    buffer: list[str] = []
    with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            buffer.append(" | ".join(row))
            if len(buffer) >= ROWS_PER_CHUNK:
                yield "\n".join(buffer) + "\n"
                buffer.clear()
    if buffer:
        yield "\n".join(buffer) + "\n"


def convert_to_text(file_path: str) -> Generator[str, None, None]:
    logger.info("Spreadsheet convert: %s", file_path)
    if file_path.lower().endswith(".csv"):
        yield from _csv_to_text(file_path)
    else:
        yield from _xlsx_to_text(file_path)
